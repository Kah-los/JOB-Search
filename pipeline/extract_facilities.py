#!/usr/bin/env python3
"""
Step 1 of the JOB-Search pipeline.

Reads "Healthcare organizations using Epic EHR.xlsx" and produces
data/facilities.json with, for every facility:
  - name, country, state
  - website_text   (the display text in the Website column)
  - hyperlink      (the embedded hyperlink target, if any)
  - career_url     (best usable career URL we already have / could salvage)
  - url_source      "hyperlink" | "text" | "salvaged-domain" | None
  - ats            detected applicant-tracking-system platform, if recognizable
  - needs_discovery  True when we still have no usable URL

It also salvages domains hidden in display text like "Careers (jefferson.edu)"
or "Job Listings at Shriners Children's (icims.com)" so the discovery step has
far fewer true blanks to chase.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Healthcare organizations using Epic EHR.xlsx"
OUT = ROOT / "data" / "facilities.json"

# Known ATS host fragments -> friendly platform name
ATS_PATTERNS = {
    "myworkdayjobs.com": "Workday",
    "workday.com": "Workday",
    "icims.com": "iCIMS",
    "taleo.net": "Taleo",
    "oraclecloud.com": "Oracle",
    "brassring.com": "Kenexa/BrassRing",
    "healthcaresource.com": "HealthcareSource",
    "selectminds.com": "SelectMinds/Oracle",
    "adp.com": "ADP",
    "workforcenow.adp.com": "ADP",
    "ultipro.com": "UKG/UltiPro",
    "ukg.com": "UKG",
    "smartrecruiters.com": "SmartRecruiters",
    "greenhouse.io": "Greenhouse",
    "lever.co": "Lever",
    "jobvite.com": "Jobvite",
    "successfactors.com": "SuccessFactors",
    "phenompeople.com": "Phenom",
    "avature.net": "Avature",
    "paylocity.com": "Paylocity",
    "applicantpro.com": "ApplicantPro",
}

URL_RE = re.compile(r"https?://[^\s)]+", re.I)
# domain inside parentheses, e.g. "(jefferson.edu)", "(icims.com)"
PAREN_DOMAIN_RE = re.compile(r"\(([a-z0-9.\-]+\.[a-z]{2,})\)", re.I)
BARE_DOMAIN_RE = re.compile(r"\b([a-z0-9\-]+(?:\.[a-z0-9\-]+)+\.(?:com|org|net|edu|gov))\b", re.I)


def is_url(s):
    return isinstance(s, str) and s.strip().lower().startswith(("http://", "https://"))


def detect_ats(url):
    if not url:
        return None
    low = url.lower()
    for frag, name in ATS_PATTERNS.items():
        if frag in low:
            return name
    return None


def salvage_domain(text):
    """Pull a probable host out of display text. Returns a host or None."""
    if not isinstance(text, str):
        return None
    m = PAREN_DOMAIN_RE.search(text)
    if m:
        return m.group(1).lower()
    m = BARE_DOMAIN_RE.search(text)
    if m:
        # avoid catching things like "U.S." — handled by regex requiring TLD
        return m.group(1).lower()
    return None


def main():
    if not XLSX.exists():
        sys.exit(f"Excel not found: {XLSX}")
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["Sheet1"]

    facilities = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if name is None or str(name).strip() == "":
            continue
        name = str(name).strip()
        country = ws.cell(r, 2).value
        state = ws.cell(r, 3).value
        wcell = ws.cell(r, 5)
        text = wcell.value if isinstance(wcell.value, str) else None
        hyperlink = wcell.hyperlink.target if wcell.hyperlink else None

        career_url = None
        url_source = None
        if is_url(hyperlink):
            career_url, url_source = hyperlink.strip(), "hyperlink"
        elif is_url(text):
            career_url, url_source = text.strip(), "text"
        else:
            host = salvage_domain(text)
            if host:
                career_url, url_source = f"https://{host}", "salvaged-domain"

        facilities.append({
            "id": len(facilities) + 1,
            "name": name,
            "country": (str(country).strip() if country else None),
            "state": (str(state).strip() if state else None),
            "website_text": text,
            "hyperlink": hyperlink,
            "career_url": career_url,
            "url_source": url_source,
            "ats": detect_ats(career_url),
            "needs_discovery": career_url is None,
        })

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(facilities, indent=2))

    total = len(facilities)
    by_source = {}
    for f in facilities:
        by_source[f["url_source"]] = by_source.get(f["url_source"], 0) + 1
    need = sum(1 for f in facilities if f["needs_discovery"])
    ats_counts = {}
    for f in facilities:
        if f["ats"]:
            ats_counts[f["ats"]] = ats_counts.get(f["ats"], 0) + 1

    print(f"Wrote {OUT}")
    print(f"Total facilities: {total}")
    print("URL source breakdown:")
    for k, v in sorted(by_source.items(), key=lambda x: -x[1]):
        print(f"  {str(k):16} {v}")
    print(f"Still needing discovery (no usable URL): {need}")
    print("Detected ATS platforms among resolved URLs:")
    for k, v in sorted(ats_counts.items(), key=lambda x: -x[1]):
        print(f"  {k:20} {v}")


if __name__ == "__main__":
    main()
